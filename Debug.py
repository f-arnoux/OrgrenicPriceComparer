import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import csv
import json
import re
from openpyxl.utils import quote_sheetname
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver import Firefox


class ProductComparer:
    def __init__(self, product_name, lafourche_site, lafourche_tag, biocoop_site, biocoop_tag, satoriz_site, satoriz_tag, greenweez_site, greenweez_tag):
        self.product_name = product_name
        self.lafourche_site = lafourche_site
        self.lafourche_tag = lafourche_tag
        self.biocoop_site = biocoop_site
        self.biocoop_tag = biocoop_tag
        self.satoriz_site = satoriz_site
        self.satoriz_tag = satoriz_tag
        self.greenweez_site = greenweez_site
        self.greenweez_tag = greenweez_tag

    def get_prices(self):
        prices = []

        # La Fourche
        lafourche_price = self._get_price_from_lafourche(self.lafourche_site)
        prices.append(lafourche_price)

        # Biocoop
        biocoop_price = self._get_price_from_site(self.biocoop_site, self.biocoop_tag)
        prices.append(biocoop_price)

        # Satoriz
        satoriz_price = self._get_price_from_site(self.satoriz_site, self.satoriz_tag)
        prices.append(satoriz_price)

        # greenweez
        greenweez_price = self._get_price_from_greenweez(self.greenweez_site, self.greenweez_tag)
        prices.append(greenweez_price)

        return prices

    def _get_price_from_site(self, url, tag):
        if url:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
            response = requests.get(url, headers=headers)
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            price_element = soup.find(class_=tag)
            price = float(re.sub(r'[^\d.,]', '', price_element.text.strip().replace(',', '.'))) if price_element else 888888
        else: price = 888888
        return price

    def _get_price_from_greenweez(self, url, tag):
        if url:
            # Spécifier le chemin complet vers le chromedriver ici
            chromedriver_path = "C:\\Users\\farnoux\\PycharmProjects\\chromedriver-win64\\chromedriver.exe"

            # Configurer les options du navigateur Chrome
            firefox_options = FirefoxOptions()
            firefox_options.add_argument("--headless")

            firefox_options.add_argument(
                'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3')
            firefox_options.add_argument('--disable-gpu')
            firefox_options.add_argument('--disable-extensions')
            firefox_options.add_argument("--disable-cache")


            # Initialiser le navigateur Chrome avec le chemin spécifié
            driver = webdriver.Chrome(options=firefox_options)

            driver.delete_all_cookies()

            # Accéder à l'URL avec le navigateur Chrome
            driver.get(url)

            # Attendre quelques secondes pour que la page se charge complètement (vous pouvez ajuster le temps d'attente selon votre besoin)
            driver.implicitly_wait(5)

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)  # Attendre que le chargement dynamique soit effectué

            # Récupérer toutes les poignées de fenêtre
            handles = driver.window_handles
            # Changer à la nouvelle fenêtre
            driver.switch_to.window(handles[0])

            # Obtenir le contenu de la page après que JavaScript ait pu modifier le DOM
            page_source = driver.page_source

            # Fermer le navigateur
            driver.quit()

            # Utiliser BeautifulSoup pour extraire les informations nécessaires
            soup = BeautifulSoup(page_source, 'html.parser')
            price_element = soup.find(class_=tag)
            price = float(re.sub(r'[^\d.,]', '', price_element.text.strip().replace(',', '.'))) if price_element else 888888

            # Afficher le prix
            print(price)
        else: price = 888888
        return price

    def _get_price_from_lafourche(self, url):
        if url:
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            text_element = soup.find('script', id='__NEXT_DATA__')
            if text_element:
                text = text_element.text.strip()
                price = json.loads(text)['props']['pageProps']['product']['meta']['finance']['unit_price']
            else:
                price = 888888
        else: price = 888888
        return price


# Liste des produits à partir du fichier CSV
products_info = []
with open('C:\\Users\\farnoux\\Documents\\liste produits.csv', 'r') as csv_file:
    csv_reader = csv.DictReader(csv_file, delimiter=',')
    for row in csv_reader:
        products_info.append({
            'Catégorie principale': row['Catégorie principale'],
            'Catégorie': row['Catégorie'],
            'Sous Catégories': row['Sous Catégories'],
            'name': row['Produit'],
            'lafourche_site': row['Site La Fourche'],
            'lafourche_tag': row['Tag La Fourche'],
            'biocoop_site': row['Site Biocoop'],
            'biocoop_tag': row['Tag Biocoop'],
            'satoriz_site': row['Site Satoriz'],
            'satoriz_tag': row['Tag Satoriz'],
            'greenweez_site': row['Site GreenWeez'],
            'greenweez_tag': row['Tag GreenWeez']
        })


# Créer un classeur Excel
workbook = Workbook()
sheet = workbook.active

# Entêtes du tableau
sheet.cell(row=1, column=1).value = 'Produit'
sheet.cell(row=1, column=2).value = 'La Fourche'
sheet.cell(row=1, column=3).value = 'Biocoop'
sheet.cell(row=1, column=4).value = 'Satoriz'
sheet.cell(row=1, column=5).value = 'greenweez'

# Obtenir les prix pour chaque produit
for row, product_info in enumerate(products_info, start=2):
    product = ProductComparer(
        product_name=product_info['name'],
        lafourche_site=product_info['lafourche_site'],
        lafourche_tag=product_info['lafourche_tag'],
        biocoop_site=product_info['biocoop_site'],
        biocoop_tag=product_info['biocoop_tag'],
        satoriz_site=product_info['satoriz_site'],
        satoriz_tag=product_info['satoriz_tag'],
        greenweez_site=product_info['greenweez_site'],
        greenweez_tag=product_info['greenweez_tag']
    )
    sheet.cell(row=row, column=1).value = product.product_name  # Nom du produit
    prices = product.get_prices()
    min_price = min(prices)  # Prix minimum dans la liste des prix
    for col, price in enumerate(prices, start=2):
        cell = sheet.cell(row=row, column=col)
        cell.value = price
        site_url = None
        if col == 2:  # La Fourche
            site_url = product.lafourche_site
        elif col == 3:  # Biocoop
            site_url = product.biocoop_site
        elif col == 4:  # Satoriz
            site_url = product.satoriz_site
        elif col == 5:  # greenweez
            site_url = product.greenweez_site

        if site_url:
            hyperlink = f'=HYPERLINK("{site_url}","{price}")'
            cell.value = hyperlink
            cell.font = Font(underline="single", color="0563C1")

        if price == min_price:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Sauvegarder le classeur Excel
workbook.save('C:\\Users\\farnoux\\Documents\\comparaison_prix.xlsx')
