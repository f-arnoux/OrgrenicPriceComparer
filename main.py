import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import csv
import os
import json
import requests
from tqdm import tqdm
from datetime import datetime
from ProductComparer import Product

metabase_elefan_start = 'https://metabase.lelefan.org/public/dashboard/53c41f3f-5644-466e-935e-897e7725f6bc?rayon=&d%25C3%25A9signation='
metabase_elefan_end = '&fournisseur=&date_d%25C3%25A9but=&date_fin='

test = False
compareLafourche = True
compareBiocoopChampollion = True
compareBiocoopFontaine = True
compareSatoriz = True
compareGreenweez = True
compareElefan = True
compareLeclerc = True
siteNameList = ["LaFourche", "Biocoop Champollion", "Biocoop Fontaine", "Satoriz", "GreenWeez", "Elefan", "Leclerc"]
to_do_list = [compareLafourche, compareBiocoopChampollion, compareBiocoopFontaine,
              compareSatoriz, compareGreenweez, compareElefan,compareLeclerc]

def extract_price_from_hyperlink(cell_value):
    # Extraire le prix de l'hyperlien
    price = None
    if isinstance(cell_value, str) and cell_value.startswith('=HYPERLINK("'):
        start_pos = cell_value.rfind(';') + 1
        if start_pos == 0:
            start_pos = cell_value.rfind(',') + 1
        end_pos = cell_value.rfind(')')
        try:
            price = float(cell_value[start_pos:end_pos])
        except ValueError:
            quote_end_pos = cell_value.rfind('"')  # Trouver la position du dernier guillemet dans la formule
            quote_start_pos = cell_value.rfind('"', 0, quote_end_pos - 1) + 1  # Trouver la position du guillemet précédent
            price = float(cell_value[quote_start_pos:quote_end_pos])
    return price

def search_product_line(reference_sheet, product_name):
    # Recherche du nom de produit dans la première colonne de la feuille de référence
    row_index = None
    reference_column = reference_sheet['A']  # Première colonne de la feuille de référence

    for cell in reference_column:
        if cell.value == product_name:
            # Correspondance trouvée, obtenir le prix de référence
            row_index = cell.row
            break  # Sortir de la boucle dès que la correspondance est trouvée
    return row_index

def set_row_fill(row_index, color_fill):
    # Recherche du nom de produit dans la première colonne de la feuille de référence
    col_index = 1
    while col_index < 16:
        cell = mois_annee_sheet.cell(row=row_index, column=col_index)
        cell.fill = color_fill
        col_index += 1


# Liste des produits à partir du fichier CSV
products_info = []
productPath = os.getcwd() + '\\liste produits.csv'
if test:
    productPath = os.getcwd() + '\\liste produits_test.csv'
with open(productPath, 'r') as csv_file:
    csv_reader = csv.DictReader(csv_file, delimiter=';')
    for row in csv_reader:
        products_info.append({
            'Catégorie principale': row['Catégorie principale'],
            'Catégorie': row['Catégorie'],
            'Sous Catégories': row['Sous Catégories'],
            'type': row['Type'],
            'name': row['Produit'],
            'lafourche_site': row['Site La Fourche'],
            'lafourche_qtt': row['Quantité La Fourche'],
            'lafourche_ean': row['Code Barre La Fourche'],
            'biocoop_champollion_site': row['Site Biocoop Champollion'],
            'biocoop_champollion_qtt': row['Quantité Biocoop Champollion'],
            'biocoop_fontaine_site': row['Site Biocoop Fontaine'],
            'biocoop_fontaine_qtt': row['Quantité Biocoop Fontaine'],
            'satoriz_site': row['Site Satoriz'],
            'satoriz_qtt': row['Quantité Satoriz'],
            'greenweez_site': row['Site GreenWeez'],
            'greenweez_qtt': row['Quantité GreenWeez'],
            'elefan_code': row['Code Elefan'],
            'elefan_qtt': row['Quantité Elefan'],
            'leclerc_site': row['Site Leclerc'],
            'leclerc_qtt': row['Quantité Leclerc'],
            'proportion': float(row['Proportion'].replace(',','.')) if row['Proportion'] else 0
        })

# Trier les produits selon les trois niveaux de catégories
sorted_products_info = sorted(products_info, key=lambda x: (x['Catégorie principale'], x['Catégorie'], x['Sous Catégories']))

# Obtenir le nom du mois et de l'année actuelle
current_date = datetime.now()
month_year = current_date.strftime('%B %Y')

# Charger la feuille "Reference 2025" pour obtenir les prix de référence
reference_workbook = openpyxl.load_workbook(os.getcwd() + '\\comparaison_prix.xlsx')
reference_sheet = reference_workbook['Reference 2025']

# Ajouter une nouvelle feuille "Mois année" si elle n'existe pas
if month_year not in reference_workbook.sheetnames:
    reference_workbook.create_sheet(month_year)

mois_annee_sheet = reference_workbook[month_year]

# Entêtes du tableau
mois_annee_sheet.cell(row=1, column=1).value = 'Produit'
mois_annee_sheet.cell(row=1, column=2).value = 'La Fourche'
mois_annee_sheet.cell(row=1, column=3).value = 'Évolution La Fourche'
mois_annee_sheet.cell(row=1, column=4).value = 'Biocoop Champollion'
mois_annee_sheet.cell(row=1, column=5).value = 'Évolution Biocoop Champollion'
mois_annee_sheet.cell(row=1, column=6).value = 'Biocoop Fontaine'
mois_annee_sheet.cell(row=1, column=7).value = 'Évolution Biocoop Fontaine'
mois_annee_sheet.cell(row=1, column=8).value = 'Satoriz'
mois_annee_sheet.cell(row=1, column=9).value = 'Évolution Satoriz'
mois_annee_sheet.cell(row=1, column=10).value = 'GreenWeez'
mois_annee_sheet.cell(row=1, column=11).value = 'Évolution GreenWeez'
mois_annee_sheet.cell(row=1, column=12).value = 'Elefan'
mois_annee_sheet.cell(row=1, column=13).value = 'Évolution Elefan'
mois_annee_sheet.cell(row=1, column=14).value = 'Leclerc'
mois_annee_sheet.cell(row=1, column=15).value = 'Évolution Leclerc'


# Variables pour le formatage
category_main_style = Font(bold=True, size=14)
category_sub_style = Font(bold=True)
separator_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
category_main_fill = PatternFill(start_color="B7C9E2", end_color="B7C9E2", fill_type="solid")  # Couleur de fond pour les catégories principales
category_sub_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Couleur de fond pour les sous-catégories

# Initialisation des totaux par site
total_nb_list = [0,0,0,0,0,0,0]

#recuperation des données de l'Elefan
response = requests.get('https://produits.lelefan.org/api/articles')
# Parse le JSON
all_elefan_product_data = response.json()  # Convertit la réponse en objet Python

# Obtenir les prix pour chaque produit
current_category_main = None
current_category_sub = None
current_row = 2  # Variable pour garder le numéro de ligne actuel
product_list = []
with tqdm(sorted_products_info, desc="Traitement des produits", dynamic_ncols=True) as pbar:
    for product_info in pbar:
        product_name = product_info['name']
        pbar.set_description(f"Traitement - {product_name}")
        # Vérifier si la catégorie principale a changé
        if current_category_main != product_info['Catégorie principale']:
            current_row += 1
            current_category_main = product_info['Catégorie principale']
            # Écrire le nom de la catégorie principale avec un style différent
            main_category_cell = mois_annee_sheet.cell(row=current_row, column=1)
            main_category_cell.value = current_category_main
            main_category_cell.font = category_main_style
            set_row_fill(current_row, category_main_fill)
            current_row += 1

        # Vérifier si la catégorie a changé
        if current_category_sub != product_info['Catégorie']:
            current_category_sub = product_info['Catégorie']
            # Écrire le nom de la catégorie avec un style différent
            sub_category_cell = mois_annee_sheet.cell(row=current_row, column=1)
            sub_category_cell.value = current_category_sub
            sub_category_cell.font = category_sub_style
            set_row_fill(current_row, category_sub_fill)
            # sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

        # Obtenir les prix pour chaque produit
        product = Product(
                product_info=product_info,
                product_list=product_list,
                elefan_data=all_elefan_product_data,
                to_do_list=to_do_list
        )

        # Écrire le nom du produit
        mois_annee_sheet.cell(row=current_row, column=1).value = product.product_name

        product_list.append(product)
        if product_info['proportion'] != 0 and not product.hasNullPrice:
            mois_annee_sheet.cell(row=current_row, column=16).value = product_info['proportion']
        # Prix minimum dans la liste des prix
        min_data = min(product.data_list, key=lambda data: data.price)

        # Écrire les prix pour chaque site
        for id, data in enumerate(product.data_list, start=0):
            if to_do_list[id]:
                if not data.isDouble and data.price != 888888:
                    total_nb_list[id] = total_nb_list[id]  + 1
                col = id + 1  # Commencer à partir de la première colonne (colonne 1)
                price_cell = mois_annee_sheet.cell(row=current_row, column=2*col)
                price_cell.value = data.price

                site_url = None
                if col == 6 and len(data.url) > 0:  # Elefan
                    site_url = metabase_elefan_start + product_info['elefan_code'] + metabase_elefan_end
                else:
                    site_url = data.url

                if site_url:
                    hyperlink = f'=HYPERLINK("{site_url}","{data.price}")'
                    price_cell.value = hyperlink
                    price_cell.font = Font(underline="single", color="0563C1")

                if data.price == min_data.price:
                    price_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                else:
                    price_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                # Calcul de l'évolution par rapport à la feuille de référence
                reference_price = None
                new_price = 888888
                reference_column = None
                row_index = search_product_line(reference_sheet=reference_sheet, product_name=product_info['name'])
                if row_index is not None:
                    if col == 1:  # Prix La Fourche
                        reference_price = extract_price_from_hyperlink(
                            reference_sheet.cell(row=row_index, column=2).value)
                        reference_column = 3  # Colonne pour l'évolution La Fourche
                    elif col == 2:  # Prix Biocoop Champollion
                        reference_price = extract_price_from_hyperlink(
                            reference_sheet.cell(row=row_index, column=4).value)
                        reference_column = 5  # Colonne pour l'évolution Biocoop Champollion
                    elif col == 3:  # Prix Biocoop Fontaine
                        reference_price = extract_price_from_hyperlink(
                            reference_sheet.cell(row=row_index, column=6).value)
                        reference_column = 7  # Colonne pour l'évolution Biocoop Fontaine
                    elif col == 4:  # Prix Satoriz
                        reference_price = extract_price_from_hyperlink(
                            reference_sheet.cell(row=row_index, column=8).value)
                        reference_column = 9  # Colonne pour l'évolution Satoriz
                    elif col == 5:  # Prix GreenWeez
                        reference_price = extract_price_from_hyperlink(
                            reference_sheet.cell(row=row_index, column=10).value)
                        reference_column = 11  # Colonne pour l'évolution GreenWeez
                        if data.price == 888888 and False: #desactive ce code
                            all_sheet = reference_workbook.sheetnames
                            index = 2
                            while (index < len(all_sheet) - 1 and new_price == 888888):
                                previous_sheet = all_sheet[len(all_sheet)-index]
                                row_index = search_product_line(reference_sheet=previous_sheet,
                                                                product_name=product_info['name'])
                                new_price = extract_price_from_hyperlink(
                                previous_sheet.cell(row=row_index, column=10).value)
                                index = index + 1
                    elif col == 6:  # Prix Elefan
                        reference_price = extract_price_from_hyperlink(
                            reference_sheet.cell(row=row_index, column=12).value)
                        reference_column = 13  # Colonne pour l'évolution Elefan
                    elif col == 7:  # Prix Leclerc
                        reference_price = extract_price_from_hyperlink(
                            reference_sheet.cell(row=row_index, column=14).value)
                        reference_column = 15  # Colonne pour l'évolution Leclerc

                if reference_price is not None:
                    price_change = ((data.price - reference_price) / reference_price) * 100
                    evolution_cell = mois_annee_sheet.cell(row=current_row, column=reference_column)
                    evolution_cell.value = f"{price_change:.2f}%"

                    if reference_price == 888888:
                        evolution_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                                                          fill_type="solid")  # Couleur de fond blanc
                    elif data.price == 888888:
                        evolution_cell.value = "-"
                        evolution_cell.fill = PatternFill(start_color="5E6D6D", end_color="5E6D6D",
                                                          fill_type="solid")  # Couleur de fond grise
                        if new_price != 888888:
                            price_cell.value = new_price

                    elif price_change < 0:
                        evolution_cell.fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD",
                                                          fill_type="solid")  # Couleur de fond bleue pastel
                    elif price_change > 0:
                        evolution_cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC",
                                                          fill_type="solid")  # Couleur de fond rouge pastel
        current_row += 1

current_row += 1
# Afficher les résultats à la fin du tableau
mois_annee_sheet.cell(row=current_row + 1, column=1).value = 'Prix du Panier'
mois_annee_sheet.cell(row=current_row + 1, column=2).value = '=SUMPRODUCT($P5:$P' + str(current_row) + '*B5:B' + str(current_row) + ')'
mois_annee_sheet.cell(row=current_row + 1, column=4).value = '=SUMPRODUCT($P5:$P' + str(current_row) + '*D5:D' + str(current_row) + ')'
mois_annee_sheet.cell(row=current_row + 1, column=6).value = '=SUMPRODUCT($P5:$P' + str(current_row) + '*F5:F' + str(current_row) + ')'
mois_annee_sheet.cell(row=current_row + 1, column=8).value = '=SUMPRODUCT($P5:$P' + str(current_row) + '*H5:H' + str(current_row) + ')'
mois_annee_sheet.cell(row=current_row + 1, column=10).value = '=SUMPRODUCT($P5:$P' + str(current_row) + '*J5:J' + str(current_row) + ')'
mois_annee_sheet.cell(row=current_row + 1, column=12).value = '=SUMPRODUCT($P5:$P' + str(current_row) + '*L5:L' + str(current_row) + ')'
mois_annee_sheet.cell(row=current_row + 1, column=14).value = '=SUMPRODUCT($P5:$P' + str(current_row) + '*N5:N' + str(current_row) + ')'

for id, data in enumerate(total_nb_list, start=0):
    print(siteNameList[id] + " : " + str(data))

# Sauvegarder le classeur Excel
reference_workbook.save(os.getcwd() + '\\comparaison_prix.xlsx')

# Ouvrir le fichier Excel après sa génération
os.startfile(os.getcwd() + '\\comparaison_prix.xlsx')